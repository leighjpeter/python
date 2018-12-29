#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pymysql as pms;
import openpyxl;
import datetime;
import smtplib;
from email.mime.text import MIMEText;
from email.mime.multipart import MIMEMultipart;
from email.header import Header;
from email.utils import parseaddr, formataddr


def get_datas(sql):
	conn = pms.connect(host='', user='', passwd='', port=, database='', charset='utf8')
	cur = conn.cursor()
	cur.execute(sql)
	datas = cur.fetchall()

	cur.close()

	return datas

def get_fields(sql):
	conn = pms.connect(host='', user='', passwd='', port=, database='', charset='utf8')
	cur = conn.cursor()
	cur.execute(sql)
	# 获取所需要的字段名
	fields = cur.description
	cur.close()

	return fields

def get_user_info(user_ids):
	conn = pms.connect(host='', user='', passwd='', port=, database='', charset='utf8')
	formater_strings = ','.join(['%s'] * len(user_ids))
	sql = "SELECT user_id as UID,real_name as 姓名,user_mobile as 手机号,identification_card as 身份证, province_name as 省, city_name as 市, 0 as 业绩 from gss_users where user_id in (%s);" % formater_strings
	cur = conn.cursor()
	cur.execute(sql, user_ids)
	user_info = cur.fetchall()
	fields = cur.description
	cur.close()

	return user_info,fields

def get_excel(data, field, file):
	new = openpyxl.Workbook()

	sheet = new.active

	sheet.title = '数据展示'
	#row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
	for col in range(len(field)):
		_ = sheet.cell(row=1, column=col+1, value=u'%s' %field[col][0])

	for row in range(len(data)):
		for col in range(len(field)):
			_ = sheet.cell(row=row+2, column=col+1, value = u'%s' %data[row][col])
	newworkbook = new.save(file)
	return newworkbook

def getYesterday():
	today = datetime.date.today()
	oneday = datetime.timedelta(days=1)
	yesterday = today - oneday
	yesterdaystr = yesterday.strftime('%Y-%m-%d')

	return yesterdaystr

def _format_addr(s):
	name, addr = parseaddr(s)
	return formataddr((Header(name,'utf-8').encode(), addr))

def  create_email(email_from, email_to, email_cc, email_subject, email_text, annex_path, annex_name):
	# 先生成一个空的带附件的实例
	message = MIMEMultipart()
	message.attach(MIMEText(email_text, 'plain', 'utf-8'))
	message['From'] = _format_addr(email_from)
	message['To'] = _format_addr(email_to)
	message['Cc'] = _format_addr(email_cc)
	#邮件主题
	message['Subject'] = Header(email_subject, 'utf-8')
	#读取附件内容
	att1 = MIMEText(open(annex_path, 'rb').read(), 'base64', 'utf-8')
	att1['Content-Type'] = 'application/octet-stream'
	#生成附件名称
	att1['Content-Disposition'] = 'attachment; filename=' + annex_name
	#将附件内容插入邮件中
	message.attach(att1)

	return message

def send_email(sender, password, receiver, msg):
	try:
		server = smtplib.SMTP_SSL("smtp.exmail.qq.com",465)
		server.ehlo()
		server.login(sender,password)
		server.sendmail(sender, receiver, msg.as_string())
		print('发送成功')
		server.quit()
	except Exception:
		print(traceback.pring_exc())
		print('发送失败')

def  main():
	mysql = "SELECT \
		a.user_id,sum(core_credit+uncore_credit) as yj \
		from gss_rebate_statics a \
		where month = '201809' and a.user_identity = 1 and a.user_id > 122 \
		GROUP BY user_id \
		ORDER BY yj DESC \
		LIMIT 100;"
	my_data = get_datas(mysql)
	user_ids = [] # list
	for t in range(len(my_data)):
		user_ids.append(my_data[t][0])

	user_info,my_field = get_user_info(user_ids)

	l = []
	for x in user_info:
		tmp = list(x)
		for i in my_data:
			if x[0] == i[0]:
				tmp.pop()
				tmp.append(i[1])

		l.append(tmp)

	my_data = l
	my_data.sort(key=lambda x:x[6], reverse=True)

	# my_field = get_fields(mysql)

	# yesterdaystr = getYesterday()

	my_file_name = '100_agent_ranking'+ (datetime.datetime.now()).strftime('%Y-%m-%d-%H') +'.xlsx'

	file_path = '/Users/jessica/Documents/ewanse/数据提取/'+ my_file_name

	get_excel(my_data,my_field,file_path)

	my_sender = ''

	my_password = ''

	my_receiver = ['']
	my_cc = []

	my_email_from = 'BI报表Robot<%s>' %my_sender

	my_email_to = '运营<%s>' % (','.join(my_receiver))

	my_email_cc = 'ET<%s>' % (','.join(my_cc))

	my_email_subject= '排名数据——'+(datetime.datetime.now()).strftime('%Y-%m-%d-%H')

	my_email_text = "Dear all,\n\t附件为业绩前100排行榜（不包含招商）\n\t如有问题及时联系 \n\t祝好！"

	my_annex_path = file_path

	my_annex_name = my_file_name

	my_msg = create_email(my_email_from, my_email_to, my_email_cc, my_email_subject, my_email_text, my_annex_path, my_annex_name)

	send_email(my_sender, my_password, my_receiver + my_cc, my_msg)

if __name__ == '__main__':
	main()





